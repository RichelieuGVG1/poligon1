import openpyxl
import os

PREFIXES = ['PEK1E', 'MOW1H']

def process_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        for row in range(1, ws.max_row + 1):
            cell_b = ws.cell(row, 2)  #столбец B
            cell_c = ws.cell(row, 3)  #столбец C

            #B или C уже заполнены то пропускаем
            if cell_b.value is not None or cell_c.value is not None:
                wb.save(file_path)
                continue

            cell_g = ws.cell(row, 7)  #столбец G
            if cell_g.value is None:
                wb.save(file_path)
                continue

            #текст по переносам строк
            lines = str(cell_g.value).splitlines()
            found = False
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                #начинается ли строка с одного из префиксов
                matched_prefix = None
                for prefix in PREFIXES:
                    if line.startswith(prefix):
                        matched_prefix = prefix
                        break

                if matched_prefix is not None:
                    #удалить префикс и пробелы после него
                    line_without_prefix = line[len(matched_prefix):].strip()
                    parts = line_without_prefix.split('/')
                    if len(parts) >= 3:  #нужно минимум 3 части (первое, второе, третье)
                        first_val = parts[0].strip()   #первое значение
                        third_val = parts[2].strip()   #третье значение
                        cell_c.value = first_val
                        cell_b.value = third_val
                        found = True
                        break

      
            wb.save(file_path)

        print(f"Обработан: {os.path.basename(file_path)}")

    except Exception as e:
        print(f"Ошибка при обработке {os.path.basename(file_path)}: {e}")

def process_all_excel_files():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_files = [f for f in os.listdir(script_dir) 
                   if f.endswith('.xlsx') and not f.startswith('~')]

    if not excel_files:
        print("В папке со скриптом не найдено .xlsx файлов")
        return

    print(f"Найдено файлов: {len(excel_files)}")
    
    for file in excel_files:
        process_excel(os.path.join(script_dir, file))
    
    print("Все файлы обработаны!")

if __name__ == "__main__":
    process_all_excel_files()
